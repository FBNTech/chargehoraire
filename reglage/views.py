from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Q
from .models import (
    Section, Departement, Mention, Niveau, Classe, Grade, CategorieEnseignant, 
    Semestre, Fonction, AnneeAcademique, Salle, Creneau, SemaineCours
)
from .forms import CreneauForm, SemaineCoursForm
import openpyxl
from django.db import transaction
from accounts.organisation_utils import get_user_organisation, is_org_user


class OrganisationFilterMixin:
    """Mixin pour filtrer les données par organisation de l'utilisateur"""
    
    def get_user_organisation(self):
        return get_user_organisation(self.request.user)
    
    def is_org_user(self):
        return is_org_user(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_org_user'] = self.is_org_user()
        context['user_organisation'] = self.get_user_organisation()
        return context


class SafeDeleteView(DeleteView):
    """
    Classe de base pour toutes les vues de suppression avec transaction atomique.
    Utilise le pattern robuste avec select_for_update() pour éviter les problèmes SQLite.
    """
    def delete(self, request, *args, **kwargs):
        try:
            # Utiliser une transaction atomique pour garantir la cohérence
            with transaction.atomic():
                # Récupérer et verrouiller l'objet pour éviter les conflits concurrents
                self.object = self.get_queryset().select_for_update().get(pk=kwargs['pk'])
                success_url = self.get_success_url()
                
                # Supprimer l'objet (les contraintes FK CASCADE s'occuperont des relations)
                self.object.delete()
            
            # Message de succès (optionnel, peut être surchargé par les sous-classes)
            if hasattr(self, 'success_message'):
                messages.success(request, self.success_message)
            
            return redirect(success_url)
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la suppression : {str(e)}")
            return redirect(self.success_url)


def gestion_entites(request):
    """Affiche la page de gestion des entités."""
    return render(request, 'reglage/gestion_entites.html')

# Vues pour Section
class SectionListView(OrganisationFilterMixin, ListView):
    model = Section
    template_name = 'reglage/section_list.html'
    context_object_name = 'sections'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        org = self.get_user_organisation()
        if org:
            # Filtrer pour ne montrer que la section de l'organisation
            return queryset.filter(CodeSection=org.code)
        return queryset

class SectionCreateView(CreateView):
    model = Section
    template_name = 'reglage/section_form.html'
    fields = ['CodeSection', 'DesignationSection']
    success_url = reverse_lazy('reglage:section_list')

class SectionUpdateView(UpdateView):
    model = Section
    template_name = 'reglage/section_form.html'
    fields = ['CodeSection', 'DesignationSection']
    success_url = reverse_lazy('reglage:section_list')

class SectionDeleteView(SafeDeleteView):
    model = Section
    template_name = 'reglage/section_confirm_delete.html'
    success_url = reverse_lazy('reglage:section_list')

# Vues pour Departement
class DepartementListView(OrganisationFilterMixin, ListView):
    model = Departement
    template_name = 'reglage/departement_list.html'
    context_object_name = 'departements'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        org = self.get_user_organisation()
        if org:
            # Filtrer par la section de l'organisation
            return queryset.filter(section__CodeSection=org.code)
        return queryset

class DepartementCreateView(CreateView):
    model = Departement
    template_name = 'reglage/departement_form.html'
    fields = ['CodeDept', 'DesignationDept', 'section']
    success_url = reverse_lazy('reglage:departement_list')

class DepartementUpdateView(UpdateView):
    model = Departement
    template_name = 'reglage/departement_form.html'
    fields = ['CodeDept', 'DesignationDept', 'section']
    success_url = reverse_lazy('reglage:departement_list')

class DepartementDeleteView(SafeDeleteView):
    model = Departement
    template_name = 'reglage/departement_confirm_delete.html'
    success_url = reverse_lazy('reglage:departement_list')

# Vues pour Mention
class MentionListView(OrganisationFilterMixin, ListView):
    model = Mention
    template_name = 'reglage/mention_list.html'
    context_object_name = 'mentions'
    
    def get_queryset(self):
        queryset = Mention.objects.all()
        # Filtrer par organisation (via departement -> section)
        user_org = self.get_user_organisation()
        if user_org:
            queryset = queryset.filter(departement__section__CodeSection=user_org.code)
        return queryset.order_by('CodeMention')

class MentionCreateView(OrganisationFilterMixin, CreateView):
    model = Mention
    template_name = 'reglage/mention_form.html'
    fields = ['CodeMention', 'DesignationMention', 'departement']
    success_url = reverse_lazy('reglage:mention_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Filtrer les départements par organisation
        user_org = self.get_user_organisation()
        if user_org:
            context['departements'] = Departement.objects.filter(section__CodeSection=user_org.code).order_by('CodeDept')
        else:
            context['departements'] = Departement.objects.all().order_by('CodeDept')
        return context

class MentionUpdateView(OrganisationFilterMixin, UpdateView):
    model = Mention
    template_name = 'reglage/mention_form.html'
    fields = ['CodeMention', 'DesignationMention', 'departement']
    success_url = reverse_lazy('reglage:mention_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Filtrer les départements par organisation
        user_org = self.get_user_organisation()
        if user_org:
            context['departements'] = Departement.objects.filter(section__CodeSection=user_org.code).order_by('CodeDept')
        else:
            context['departements'] = Departement.objects.all().order_by('CodeDept')
        return context

class MentionDeleteView(SafeDeleteView):
    model = Mention
    template_name = 'reglage/mention_confirm_delete.html'
    success_url = reverse_lazy('reglage:mention_list')

# Vues pour Niveau
class NiveauListView(ListView):
    model = Niveau
    template_name = 'reglage/niveau_list.html'
    context_object_name = 'niveaux'

class NiveauCreateView(CreateView):
    model = Niveau
    template_name = 'reglage/niveau_form.html'
    fields = ['CodeNiveau', 'DesignationNiveau']
    success_url = reverse_lazy('reglage:niveau_list')

class NiveauUpdateView(UpdateView):
    model = Niveau
    template_name = 'reglage/niveau_form.html'
    fields = ['CodeNiveau', 'DesignationNiveau']
    success_url = reverse_lazy('reglage:niveau_list')

class NiveauDeleteView(SafeDeleteView):
    model = Niveau
    template_name = 'reglage/niveau_confirm_delete.html'
    success_url = reverse_lazy('reglage:niveau_list')

# Vues pour Classe
class ClasseListView(OrganisationFilterMixin, ListView):
    model = Classe
    template_name = 'reglage/classe_list.html'
    context_object_name = 'classes'
    
    def get_queryset(self):
        queryset = Classe.objects.all()
        # Filtrer par organisation (via mention -> departement -> section)
        user_org = self.get_user_organisation()
        if user_org:
            queryset = queryset.filter(mention__departement__section__CodeSection=user_org.code)
        return queryset.order_by('CodeClasse')

class ClasseCreateView(OrganisationFilterMixin, CreateView):
    model = Classe
    template_name = 'reglage/classe_form.html'
    fields = ['niveau', 'mention']
    success_url = reverse_lazy('reglage:classe_list')
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Filtrer les mentions par organisation
        user_org = self.get_user_organisation()
        if user_org:
            form.fields['mention'].queryset = Mention.objects.filter(
                departement__section__CodeSection=user_org.code
            ).order_by('CodeMention')
        return form

class ClasseUpdateView(OrganisationFilterMixin, UpdateView):
    model = Classe
    template_name = 'reglage/classe_form.html'
    fields = ['niveau', 'mention']
    success_url = reverse_lazy('reglage:classe_list')
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Filtrer les mentions par organisation
        user_org = self.get_user_organisation()
        if user_org:
            form.fields['mention'].queryset = Mention.objects.filter(
                departement__section__CodeSection=user_org.code
            ).order_by('CodeMention')
        return form

class ClasseDeleteView(SafeDeleteView):
    model = Classe
    template_name = 'reglage/classe_confirm_delete.html'
    success_url = reverse_lazy('reglage:classe_list')

def classe_import_excel(request):
    """Importer des classes depuis un fichier Excel"""
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('excel_file')
            replace_all = request.POST.get('replace_all') == 'on'
            
            if not excel_file:
                messages.error(request, 'Aucun fichier sélectionné')
                return redirect('reglage:classe_list')
            
            # Vérifier l'extension du fichier
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Format de fichier invalide. Utilisez .xlsx ou .xls')
                return redirect('reglage:classe_list')
            
            # Charger le fichier Excel
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active
            except Exception as e:
                messages.error(request, f'Erreur lors de la lecture du fichier Excel: {str(e)}')
                return redirect('reglage:classe_list')
            
            # Compter les lignes (en excluant l'en-tête)
            total_rows = sheet.max_row - 1
            
            if total_rows < 1:
                messages.warning(request, 'Le fichier Excel est vide')
                return redirect('reglage:classe_list')
            
            classes_created = 0
            classes_updated = 0
            errors = []
            
            with transaction.atomic():
                # Supprimer toutes les classes si demandé
                if replace_all:
                    deleted_count = Classe.objects.all().delete()[0]
                    messages.info(request, f'{deleted_count} classe(s) supprimée(s)')
                
                # Parcourir les lignes (en sautant l'en-tête)
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Vérifier que la ligne contient des données
                        if not any(row):
                            continue
                        
                        code_classe = str(row[0]).strip() if row[0] else None
                        designation = str(row[1]).strip() if row[1] else None
                        niveau = str(row[2]).strip() if row[2] else None
                        mention = str(row[3]).strip() if row[3] else None
                        
                        # Validation
                        if not code_classe:
                            errors.append(f"Ligne {row_num}: CodeClasse manquant")
                            continue
                        
                        if not designation:
                            errors.append(f"Ligne {row_num}: DesignationClasse manquant")
                            continue
                        
                        if not niveau:
                            errors.append(f"Ligne {row_num}: Niveau manquant")
                            continue
                        
                        # Récupérer ou créer les objets Niveau et Mention
                        try:
                            # Tenter par désignation
                            niveau_obj = Niveau.objects.filter(DesignationNiveau=niveau).first()
                            if not niveau_obj:
                                # Tenter par code (si un code identique existe déjà)
                                niveau_obj = Niveau.objects.filter(CodeNiveau=niveau).first()
                                if niveau_obj:
                                    # Mettre à jour la désignation si différente
                                    if niveau_obj.DesignationNiveau != niveau:
                                        niveau_obj.DesignationNiveau = niveau
                                        niveau_obj.save(update_fields=["DesignationNiveau"]) 
                                else:
                                    niveau_obj = Niveau.objects.create(CodeNiveau=niveau, DesignationNiveau=niveau)
                        except Exception as e:
                            errors.append(f"Ligne {row_num}: Erreur niveau - {str(e)}")
                            continue
                        
                        mention_obj = None
                        if mention:
                            try:
                                mention_obj = Mention.objects.filter(DesignationMention=mention).first()
                                if not mention_obj:
                                    mention_obj = Mention.objects.filter(CodeMention=mention).first()
                                    if mention_obj:
                                        if mention_obj.DesignationMention != mention:
                                            mention_obj.DesignationMention = mention
                                            mention_obj.save(update_fields=["DesignationMention"]) 
                                    else:
                                        mention_obj = Mention.objects.create(CodeMention=mention, DesignationMention=mention)
                            except Exception as e:
                                errors.append(f"Ligne {row_num}: Erreur mention - {str(e)}")
                                continue
                        
                        # Créer ou mettre à jour la classe
                        classe, created = Classe.objects.update_or_create(
                            CodeClasse=code_classe,
                            defaults={
                                'DesignationClasse': designation,
                                'niveau': niveau_obj,
                                'mention': mention_obj
                            }
                        )
                        
                        if created:
                            classes_created += 1
                        else:
                            classes_updated += 1
                            
                    except Exception as e:
                        errors.append(f"Ligne {row_num}: {str(e)}")
                        continue
            
            # Messages de résultat
            if classes_created > 0:
                messages.success(request, f'✅ {classes_created} classe(s) créée(s) avec succès')
            
            if classes_updated > 0:
                messages.info(request, f'ℹ️ {classes_updated} classe(s) mise(s) à jour')
            
            if errors:
                messages.warning(request, f'⚠️ {len(errors)} erreur(s) rencontrée(s):')
                for error in errors[:10]:  # Limiter à 10 erreurs affichées
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.error(request, f'... et {len(errors) - 10} autre(s) erreur(s)')
            
            if classes_created == 0 and classes_updated == 0:
                messages.warning(request, 'Aucune classe importée')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'importation: {str(e)}')
        
        return redirect('reglage:classe_list')
    
    return redirect('reglage:classe_list')

# Vues pour Grade
class GradeListView(ListView):
    model = Grade
    template_name = 'reglage/grade_list.html'
    context_object_name = 'grades'

class GradeCreateView(CreateView):
    model = Grade
    template_name = 'reglage/grade_form.html'
    fields = ['CodeGrade', 'DesignationGrade']
    success_url = reverse_lazy('reglage:grade_list')

class GradeUpdateView(UpdateView):
    model = Grade
    template_name = 'reglage/grade_form.html'
    fields = ['CodeGrade', 'DesignationGrade']
    success_url = reverse_lazy('reglage:grade_list')

class GradeDeleteView(SafeDeleteView):
    model = Grade
    template_name = 'reglage/grade_confirm_delete.html'
    success_url = reverse_lazy('reglage:grade_list')

# Vues pour CategorieEnseignant
class CategorieListView(ListView):
    model = CategorieEnseignant
    template_name = 'reglage/categorie_list.html'
    context_object_name = 'categories'

class CategorieCreateView(CreateView):
    model = CategorieEnseignant
    template_name = 'reglage/categorie_form.html'
    fields = ['CodeCategorie', 'DesignationCategorie']
    success_url = reverse_lazy('reglage:categorie_list')

class CategorieUpdateView(UpdateView):
    model = CategorieEnseignant
    template_name = 'reglage/categorie_form.html'
    fields = ['CodeCategorie', 'DesignationCategorie']
    success_url = reverse_lazy('reglage:categorie_list')

class CategorieDeleteView(SafeDeleteView):
    model = CategorieEnseignant
    template_name = 'reglage/categorie_confirm_delete.html'
    success_url = reverse_lazy('reglage:categorie_list')

# Vues pour Semestre
class SemestreListView(ListView):
    model = Semestre
    template_name = 'reglage/semestre_list.html'
    context_object_name = 'semestres'

class SemestreCreateView(CreateView):
    model = Semestre
    template_name = 'reglage/semestre_form.html'
    fields = ['CodeSemestre', 'DesignationSemestre']
    success_url = reverse_lazy('reglage:semestre_list')

class SemestreUpdateView(UpdateView):
    model = Semestre
    template_name = 'reglage/semestre_form.html'
    fields = ['CodeSemestre', 'DesignationSemestre']
    success_url = reverse_lazy('reglage:semestre_list')

class SemestreDeleteView(SafeDeleteView):
    model = Semestre
    template_name = 'reglage/semestre_confirm_delete.html'
    success_url = reverse_lazy('reglage:semestre_list')

# Vues pour Fonction
class FonctionListView(ListView):
    model = Fonction
    template_name = 'reglage/fonction_list.html'
    context_object_name = 'fonctions'

class FonctionCreateView(CreateView):
    model = Fonction
    template_name = 'reglage/fonction_form.html'
    fields = ['CodeFonction', 'DesignationFonction']
    success_url = reverse_lazy('reglage:fonction_list')

class FonctionUpdateView(UpdateView):
    model = Fonction
    template_name = 'reglage/fonction_form.html'
    fields = ['CodeFonction', 'DesignationFonction']
    success_url = reverse_lazy('reglage:fonction_list')

class FonctionDeleteView(SafeDeleteView):
    model = Fonction
    template_name = 'reglage/fonction_confirm_delete.html'
    success_url = reverse_lazy('reglage:fonction_list')


# ==================== NOUVELLES VUES ====================

# Vues pour AnneeAcademique
class AnneeAcademiqueListView(ListView):
    model = AnneeAcademique
    template_name = 'reglage/annee_list.html'
    context_object_name = 'annees'

class AnneeAcademiqueCreateView(CreateView):
    model = AnneeAcademique
    template_name = 'reglage/annee_form.html'
    fields = ['code', 'designation', 'date_debut', 'date_fin', 'est_en_cours']
    success_url = reverse_lazy('reglage:annee_list')
    
    def form_valid(self, form):
        messages.success(self.request, "Année académique créée avec succès.")
        return super().form_valid(form)

class AnneeAcademiqueUpdateView(UpdateView):
    model = AnneeAcademique
    template_name = 'reglage/annee_form.html'
    fields = ['code', 'designation', 'date_debut', 'date_fin', 'est_en_cours']
    success_url = reverse_lazy('reglage:annee_list')
    
    def form_valid(self, form):
        messages.success(self.request, "Année académique modifiée avec succès.")
        return super().form_valid(form)

class AnneeAcademiqueDeleteView(SafeDeleteView):
    model = AnneeAcademique
    template_name = 'reglage/annee_confirm_delete.html'
    success_url = reverse_lazy('reglage:annee_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, "Année académique supprimée avec succès.")
        return super().delete(request, *args, **kwargs)


# Vues pour Salle
class SalleListView(ListView):
    model = Salle
    template_name = 'reglage/salle_list.html'
    context_object_name = 'salles'
    
    def get_queryset(self):
        queryset = Salle.objects.all()
        type_filtre = self.request.GET.get('type')
        if type_filtre:
            queryset = queryset.filter(type_salle=type_filtre)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['types_salle'] = Salle._meta.get_field('type_salle').choices
        return context

class SalleCreateView(CreateView):
    model = Salle
    template_name = 'reglage/salle_form.html'
    fields = ['code', 'designation', 'capacite', 'type_salle', 'est_disponible', 'remarques']
    success_url = reverse_lazy('reglage:salle_list')
    
    def form_valid(self, form):
        messages.success(self.request, "Salle créée avec succès.")
        return super().form_valid(form)

class SalleUpdateView(UpdateView):
    model = Salle
    template_name = 'reglage/salle_form.html'
    fields = ['code', 'designation', 'capacite', 'type_salle', 'est_disponible', 'remarques']
    success_url = reverse_lazy('reglage:salle_list')
    
    def form_valid(self, form):
        messages.success(self.request, "Salle modifiée avec succès.")
        return super().form_valid(form)

class SalleDeleteView(SafeDeleteView):
    model = Salle
    template_name = 'reglage/salle_confirm_delete.html'
    success_url = reverse_lazy('reglage:salle_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, "Salle supprimée avec succès.")
        return super().delete(request, *args, **kwargs)


# Vues pour Creneau
class CreneauListView(ListView):
    model = Creneau
    template_name = 'reglage/creneau_list.html'
    context_object_name = 'creneaux'
    
    def get_queryset(self):
        """Filtrer les créneaux selon l'organisation de l'utilisateur"""
        queryset = super().get_queryset()
        
        # Si l'utilisateur est un superuser, afficher tous les créneaux
        if self.request.user.is_superuser:
            return queryset
        
        # Pour les utilisateurs d'organisations, filtrer par section de l'organisation
        if hasattr(self.request.user, 'profile') and self.request.user.profile.organisation:
            # Le code de l'organisation correspond au CodeSection
            org_code = self.request.user.profile.organisation.code
            
            # Récupérer la section correspondant à l'organisation
            from reglage.models import Section
            try:
                section_org = Section.objects.get(CodeSection=org_code)
                
                # Afficher les créneaux sans section (généraux) + créneaux de l'organisation
                queryset = queryset.filter(
                    Q(section__isnull=True) |  # Créneaux généraux (toutes les sections)
                    Q(section=section_org)  # Créneaux spécifiques à l'organisation
                )
            except Section.DoesNotExist:
                # Si la section n'existe pas, afficher uniquement les créneaux généraux
                queryset = queryset.filter(section__isnull=True)
        
        return queryset

class CreneauCreateView(CreateView):
    model = Creneau
    template_name = 'reglage/creneau_form.html'
    form_class = CreneauForm
    success_url = reverse_lazy('reglage:creneau_list')
    
    def form_valid(self, form):
        messages.success(self.request, "Créneau créé avec succès.")
        return super().form_valid(form)

class CreneauUpdateView(UpdateView):
    model = Creneau
    template_name = 'reglage/creneau_form.html'
    form_class = CreneauForm
    success_url = reverse_lazy('reglage:creneau_list')
    
    def form_valid(self, form):
        messages.success(self.request, "Créneau modifié avec succès.")
        return super().form_valid(form)

class CreneauDeleteView(SafeDeleteView):
    model = Creneau
    template_name = 'reglage/creneau_confirm_delete.html'
    success_url = reverse_lazy('reglage:creneau_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, "Créneau supprimé avec succès.")
        return super().delete(request, *args, **kwargs)


# Vues pour SemaineCours
class SemaineCoursListView(ListView):
    model = SemaineCours
    template_name = 'reglage/semaine_list.html'
    context_object_name = 'semaines'
    
    def get_queryset(self):
        queryset = SemaineCours.objects.all()
        annee = self.request.GET.get('annee')
        if annee:
            queryset = queryset.filter(annee_academique=annee)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Récupérer les années académiques pour le filtre
        context['annees'] = SemaineCours.objects.values_list(
            'annee_academique', flat=True
        ).distinct().order_by('-annee_academique')
        return context

class SemaineCoursCreateView(CreateView):
    model = SemaineCours
    template_name = 'reglage/semaine_form.html'
    form_class = SemaineCoursForm
    success_url = reverse_lazy('reglage:semaine_list')
    
    def form_valid(self, form):
        messages.success(self.request, "Semaine de cours créée avec succès.")
        return super().form_valid(form)

class SemaineCoursUpdateView(UpdateView):
    model = SemaineCours
    template_name = 'reglage/semaine_form.html'
    form_class = SemaineCoursForm
    success_url = reverse_lazy('reglage:semaine_list')
    
    def form_valid(self, form):
        messages.success(self.request, "Semaine de cours modifiée avec succès.")
        return super().form_valid(form)

class SemaineCoursDeleteView(SafeDeleteView):
    model = SemaineCours
    template_name = 'reglage/semaine_confirm_delete.html'
    success_url = reverse_lazy('reglage:semaine_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, "Semaine de cours supprimée avec succès.")
        return super().delete(request, *args, **kwargs)


# Vues de suppression en masse (superuser uniquement)
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_POST

@require_POST
def mention_delete_all(request):
    """Supprimer toutes les mentions - accessible uniquement au superuser"""
    if not request.user.is_superuser:
        messages.error(request, "Vous n'avez pas la permission d'effectuer cette action.")
        return redirect('reglage:mention_list')
    
    try:
        with transaction.atomic():
            count = Mention.objects.count()
            Mention.objects.all().delete()
            messages.success(request, f"{count} mention(s) supprimée(s) avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression : {str(e)}")
    
    return redirect('reglage:mention_list')


@require_POST
def classe_delete_all(request):
    """Supprimer toutes les classes - accessible uniquement au superuser"""
    if not request.user.is_superuser:
        messages.error(request, "Vous n'avez pas la permission d'effectuer cette action.")
        return redirect('reglage:classe_list')
    
    try:
        with transaction.atomic():
            count = Classe.objects.count()
            Classe.objects.all().delete()
            messages.success(request, f"{count} classe(s) supprimée(s) avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression : {str(e)}")
    
    return redirect('reglage:classe_list')


def mention_generate_classes(request, pk):
    """Générer automatiquement les classes pour tous les niveaux à partir d'une mention"""
    mention = get_object_or_404(Mention, pk=pk)
    niveaux = Niveau.objects.all()
    
    if not niveaux.exists():
        messages.warning(request, "Aucun niveau n'est défini. Veuillez d'abord créer des niveaux.")
        return redirect('reglage:mention_list')
    
    created_count = 0
    existing_count = 0
    
    try:
        with transaction.atomic():
            for niveau in niveaux:
                # Vérifier si la classe existe déjà
                if not Classe.objects.filter(niveau=niveau, mention=mention).exists():
                    Classe.objects.create(niveau=niveau, mention=mention)
                    created_count += 1
                else:
                    existing_count += 1
        
        if created_count > 0:
            messages.success(request, f"{created_count} classe(s) créée(s) pour la mention {mention.CodeMention}.")
        if existing_count > 0:
            messages.info(request, f"{existing_count} classe(s) existai(en)t déjà.")
        if created_count == 0 and existing_count > 0:
            messages.info(request, f"Toutes les classes existent déjà pour la mention {mention.CodeMention}.")
            
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération des classes : {str(e)}")
    
    return redirect('reglage:classe_list')


def mention_generate_all_classes(request):
    """Générer automatiquement les classes pour toutes les mentions et tous les niveaux"""
    # Filtrer les mentions par organisation de l'utilisateur
    user_org = get_user_organisation(request.user)
    if user_org:
        mentions = Mention.objects.filter(departement__section__CodeSection=user_org.code)
    else:
        mentions = Mention.objects.all()
    
    niveaux = Niveau.objects.all()
    
    if not mentions.exists():
        messages.warning(request, "Aucune mention n'est définie. Veuillez d'abord créer des mentions.")
        return redirect('reglage:mention_list')
    
    if not niveaux.exists():
        messages.warning(request, "Aucun niveau n'est défini. Veuillez d'abord créer des niveaux.")
        return redirect('reglage:mention_list')
    
    created_count = 0
    existing_count = 0
    
    try:
        with transaction.atomic():
            for mention in mentions:
                for niveau in niveaux:
                    # Vérifier si la classe existe déjà
                    if not Classe.objects.filter(niveau=niveau, mention=mention).exists():
                        Classe.objects.create(niveau=niveau, mention=mention)
                        created_count += 1
                    else:
                        existing_count += 1
        
        if created_count > 0:
            messages.success(request, f"{created_count} classe(s) créée(s) pour toutes les mentions.")
        if existing_count > 0:
            messages.info(request, f"{existing_count} classe(s) existai(en)t déjà et ont été ignorées.")
        if created_count == 0 and existing_count > 0:
            messages.info(request, "Toutes les classes existent déjà pour toutes les mentions.")
            
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération des classes : {str(e)}")
    
    return redirect('reglage:classe_list')
